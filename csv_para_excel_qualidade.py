"""
Converte cada CSV da pasta qualidade2 em um arquivo Excel separado,
com colunas formatadas e coloridas para fácil visualização.

COMO USAR:
    pip install openpyxl pandas
    python csv_para_excel_v2.py
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

PASTA_QUALIDADE = r""

# =============================================================================
# CORES
# =============================================================================

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def borda():
    lado = Side(style="thin", color="BFC9D4")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

COR_HEADER      = "1F3864"
COR_LINHA_PAR   = "EEF2F7"
COR_LINHA_IMPAR = "FFFFFF"

CORES_CLASS = {
    "SIM":     ("C6EFCE", "276221"),
    "PARCIAL": ("FFEB9C", "9C6500"),
    "NÃO":     ("FFCCCC", "9C0006"),
    "NAO":     ("FFCCCC", "9C0006"),
}
CORES_APROV = {
    "SIM": ("C6EFCE", "276221"),
    "NÃO": ("FFCCCC", "9C0006"),
    "NAO": ("FFCCCC", "9C0006"),
}
CORES_REVISAO = {
    "SIM": ("FFEB9C", "9C6500"),
    "NÃO": ("FFFFFF", "666666"),
    "NAO": ("FFFFFF", "666666"),
}

LARGURAS = {
    "arquivo":        32,
    "base":           12,
    "titulo":         45,
    "ano":             7,
    "score_total":     8,
    "aprovado":       10,
    "revisao_manual": 14,
    "motivo_revisao": 38,
    "observacoes":    35,
    "erro":           30,
    "_class":         13,
    "_just":          40,
    "_trecho":        40,
}

# 9 questões (v2)
COLUNAS_COLORIDAS_CLASS = {f"q{i}" for i in range(1, 10)}
COLUNAS_COLORIDAS_APROV = {"aprovado"}
COLUNAS_COLORIDAS_REVIS = {"revisao_manual"}

# =============================================================================
# CONVERTE UM CSV -> XLSX
# =============================================================================

def converter(csv_path: Path):
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig").fillna("")

    wb = Workbook()
    ws = wb.active
    ws.title = "Avaliacao"
    ws.sheet_view.showGridLines = False

    # Cabeçalho
    for col_idx, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = fill(COR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda()
    ws.row_dimensions[1].height = 36

    # Dados
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        bg = COR_LINHA_PAR if row_idx % 2 == 0 else COR_LINHA_IMPAR

        for col_idx, col_name in enumerate(df.columns, 1):
            valor = row_data[col_name]
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.border    = borda()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font      = Font(name="Arial", size=9)

            val_upper = str(valor).strip().upper()

            if col_name in COLUNAS_COLORIDAS_CLASS and val_upper in CORES_CLASS:
                bg_c, txt_c = CORES_CLASS[val_upper]
                c.fill      = fill(bg_c)
                c.font      = Font(name="Arial", bold=True, color=txt_c, size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")

            elif col_name in COLUNAS_COLORIDAS_APROV and val_upper in CORES_APROV:
                bg_c, txt_c = CORES_APROV[val_upper]
                c.fill      = fill(bg_c)
                c.font      = Font(name="Arial", bold=True, color=txt_c, size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")

            elif col_name in COLUNAS_COLORIDAS_REVIS and val_upper in CORES_REVISAO:
                bg_c, txt_c = CORES_REVISAO[val_upper]
                c.fill      = fill(bg_c)
                c.font      = Font(name="Arial", bold=True, color=txt_c, size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")

            elif col_name == "score_total":
                c.fill      = fill(bg)
                c.font      = Font(name="Arial", bold=True, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")

            else:
                c.fill = fill(bg)

        ws.row_dimensions[row_idx].height = 55

    # Larguras
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in LARGURAS:
            w = LARGURAS[col_name]
        elif col_name.endswith("_just") or col_name.endswith("_trecho"):
            w = LARGURAS["_just"]
        elif col_name in COLUNAS_COLORIDAS_CLASS:
            w = LARGURAS["_class"]
        else:
            w = 20
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    xlsx_path = csv_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    print(f"  OK  {csv_path.name}  ->  {xlsx_path.name}  ({len(df)} linhas)")

# =============================================================================
# PONTO DE ENTRADA
# =============================================================================

def main():
    pasta = Path(PASTA_QUALIDADE)
    csvs  = sorted(pasta.rglob("*.csv"))

    if not csvs:
        print(f"Nenhum CSV encontrado em: {pasta}")
        return

    print(f"{'='*55}")
    print(f"Convertendo {len(csvs)} arquivo(s) CSV -> Excel")
    print(f"{'='*55}")

    for csv_path in csvs:
        converter(csv_path)

    print(f"\nConcluído! {len(csvs)} arquivo(s) Excel gerado(s).")


if __name__ == "__main__":
    main()