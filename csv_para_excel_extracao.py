import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
PASTA_EXTRACAO = r""

# Estilos de cores
COR_HEADER      = "203764"  # Azul escuro
COR_LINHA_PAR   = "F2F2F2"  # Cinza claro
COR_LINHA_IMPAR = "FFFFFF"  # Branco

# Cores para o Status/Revisão
CORES_STATUS = {
    "OK":      ("C6EFCE", "006100"), # Verde
    "REVISAR": ("FFC7CE", "9C0006"), # Vermelho
}

# Larguras específicas para os campos da RSL
LARGURAS = {
    "arquivo":               30,
    "titulo":                50,
    "ano":                   8,
    "objetivo_principal":    45,
    "contexto_avaliacao":    25,
    "duracao_estudo":        15,
    "numero_participantes":  15,
    "tipo_abordagem":        15,
    "nome_metodo":           35,
    "descricao_abordagem":   50,
    "vantagens":             35,
    "desvantagens":          35,
    "metricas_coleta":       35,
    "tecnicas_analise":      35,
    "visualizacao_dados":    25,
    "trabalhos_futuros":     40,
    "status":                12,
    "motivo_revisao":        40,
    "erro":                  30
}

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def borda():
    lado = Side(style="thin", color="BFC9D4")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def converter_extracao(csv_path: Path):
    # Lê o CSV gerado pelo script anterior
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig").fillna("")

    wb = Workbook()
    ws = wb.active
    ws.title = "Extração UX"
    ws.sheet_view.showGridLines = False

    # Cabeçalho
    for col_idx, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        c.font      = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
        c.fill      = fill(COR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda()
    ws.row_dimensions[1].height = 40

    # Dados
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        bg = COR_LINHA_PAR if row_idx % 2 == 0 else COR_LINHA_IMPAR
        
        for col_idx, col_name in enumerate(df.columns, 1):
            valor = row_data[col_name]
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.border    = borda()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font      = Font(name="Segoe UI", size=9)
            c.fill      = fill(bg)

            # Colorir coluna de Status
            if col_name == "status":
                val_upper = str(valor).strip().upper()
                if val_upper in CORES_STATUS:
                    bg_s, txt_s = CORES_STATUS[val_upper]
                    c.fill = fill(bg_s)
                    c.font = Font(name="Segoe UI", bold=True, color=txt_s)
                    c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 80 # Altura maior para ler os resumos

    # Aplicar Larguras
    for col_idx, col_name in enumerate(df.columns, 1):
        largura = LARGURAS.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    # Congelar painel superior
    ws.freeze_panes = "A2"

    xlsx_path = csv_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    print(f"  ✔ Planilha gerada: {xlsx_path.name}")

def main():
    pasta = Path(PASTA_EXTRACAO)
    csvs  = sorted(pasta.glob("extracao_*.csv"))

    if not csvs:
        print(f"Nenhum CSV de extração encontrado em: {pasta}")
        return

    print("-" * 50)
    print(f"Convertendo extrações para Excel...")
    print("-" * 50)

    for csv_path in csvs:
        converter_extracao(csv_path)

    print("\nProcesso concluído! Agora você pode revisar os dados com calma.")

if __name__ == "__main__":
    main()